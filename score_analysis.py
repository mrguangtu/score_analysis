import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os

class ScoreAnalyzer:
    def __init__(self):
        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False
        
    def load_data(self, file_path):
        self.df = pd.read_excel(file_path)
        # 如果发现列名有问题，手动重命名
        if '校总分排名' in self.df.columns:
            self.df = self.df.rename(columns={'校总分排名': '校总分排名'})
        # 打印列名以验证
        print("数据框的列名：", self.df.columns.tolist())
        
    def create_scatter_plot(self, data, title, save_path=None, return_image=False):
        """创建散点图，可选择保存为文件或返回图像对象"""
        # 创建图形和轴
        fig = plt.figure(figsize=(10, 7))
        ax = plt.subplot(111)
        plt.subplots_adjust(right=0.85)

        # 获取所有班级并设置颜色
        unique_classes = data['班级'].unique()
        colors = plt.cm.rainbow(np.linspace(0, 1, len(unique_classes)))

        # 绘制散点图
        for class_num, color in zip(unique_classes, colors):
            mask = data['班级'] == class_num
            ax.scatter(data[mask]['校化学排名'], 
                      data[mask]['校总分排名'],
                      alpha=0.6,
                      color=color,
                      label=f'班级 {class_num}',
                      s=30)

        # 添加对角线
        ax.plot([0, 800], [0, 800], 'k--', alpha=0.5)

        # 添加上下边界区域
        x = np.linspace(0, 800, 100)
        # 上三角形区域
        ax.fill_between(x, x + 100, 800, alpha=0.2, color='lightblue')
        # 下三角形区域
        ax.fill_between(x, 0, x - 100, alpha=0.2, color='lightblue')

        # 设置标题和标签
        ax.set_title(title, fontsize=14, pad=15)
        ax.set_xlabel('校化学排名')
        ax.set_ylabel('校总分排名')

        # 添加网格和文字标注
        ax.grid(True, linestyle='--', alpha=0.3)
        ax.text(80, 640, '化学好，其它科拖后腿', fontsize=12)
        ax.text(480, 160, '其它科好，化学拖后腿', fontsize=12)

        # 添加图例
        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')

        # 设置坐标轴范围和比例
        ax.set_xlim(0, 800)
        ax.set_ylim(0, 800)
        ax.set_aspect('equal')

        # 保存或显示图形
        plt.tight_layout()
        if return_image:
            img_bio = BytesIO()
            plt.savefig(img_bio, format='png', dpi=300, bbox_inches='tight')
            plt.close()
            img_bio.seek(0)
            return img_bio
        elif save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            plt.close()
        else:
            plt.show()

    def get_biased_students(self, data):
        """识别偏科学生"""
        chemistry_biased = data[data['校总分排名'] > data['校化学排名'] + 100].copy()
        others_biased = data[data['校总分排名'] < data['校化学排名'] - 100].copy()
        
        return {
            '化学好其他差': chemistry_biased.sort_values('校化学排名'),
            '其他好化学差': others_biased.sort_values('校总分排名')
        }

    def generate_all_plots(self, output_dir='plots'):
        """生成所有图片并保存到指定目录"""
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        self.plot_paths = {}
        
        # 生成总体分析图
        total_plot_path = os.path.join(output_dir, 'total_analysis.png')
        self.create_scatter_plot(self.df, '化学排名与总分排名对比分析', save_path=total_plot_path)
        self.plot_paths['总体分析'] = total_plot_path
        
        # 为每个班级生成图片，使用整数类型的班级编号
        class_nums = sorted([int(x) for x in self.df['班级'].unique()])
        for class_num in class_nums:
            class_data = self.df[self.df['班级'] == class_num]
            plot_path = os.path.join(output_dir, f'class_{class_num}.png')
            self.create_scatter_plot(
                class_data,
                f'班级{class_num}化学排名与总分排名对比分析',
                save_path=plot_path
            )
            self.plot_paths[f'班级{class_num}'] = plot_path

    def get_triangle_area_students(self, data):
        """获取三角形区域内的学生数据"""
        # 获取化学好其他差的学生（上三角区域）
        upper_triangle = data[data['校总分排名'] > data['校化学排名'] + 100].copy()
        upper_triangle['区域'] = '化学好其他差'
        
        # 获取其他好化学差的学生（下三角区域）
        lower_triangle = data[data['校总分排名'] < data['校化学排名'] - 100].copy()
        lower_triangle['区域'] = '其他好化学差'
        
        # 合并两个区域的数据
        triangle_data = pd.concat([upper_triangle, lower_triangle])
        # 按总分排名排序
        triangle_data = triangle_data.sort_values('校总分排名')
        
        return triangle_data

    def export_analysis_to_excel(self, output_path='分析结果.xlsx'):
        """导出分析结果到Excel"""
        # 创建新的工作簿
        wb = Workbook()
        wb.remove(wb.active)
        
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        writer.book = wb
        
        # 创建总体分析sheet
        worksheet = writer.book.create_sheet('总体分析')
        img = Image(self.plot_paths['总体分析'])
        worksheet.add_image(img, 'A1')
        
        # 写入总体分析数据
        start_row = 120
        triangle_data = self.get_triangle_area_students(self.df)
        worksheet.cell(row=start_row, column=1, value='三角形区域内的学生数据分析')
        
        # 将数据转换为列表并直接写入worksheet
        headers = list(triangle_data.columns)
        for col_idx, header in enumerate(headers, 1):
            worksheet.cell(row=start_row + 2, column=col_idx, value=header)
        
        for row_idx, row in enumerate(triangle_data.values, 1):
            for col_idx, value in enumerate(row, 1):
                worksheet.cell(row=start_row + 2 + row_idx, column=col_idx, value=value)
        
        # 为每个班级创建sheet，使用整数排序确保正确顺序
        class_nums = sorted([int(x) for x in self.df['班级'].unique()])  # 转换为整数后排序
        
        for class_num in class_nums:
            # 使用整数进行数据筛选
            class_data = self.df[self.df['班级'] == class_num]
            sheet_name = f'班级{class_num}'
            
            # 创建工作表
            worksheet = writer.book.create_sheet(sheet_name)
            
            # 添加图片
            img = Image(self.plot_paths[sheet_name])
            worksheet.add_image(img, 'A1')
            
            # 写入数据
            start_row = 120
            
            # 获取该班级的三角形区域数据
            upper_triangle = class_data[class_data['校总分排名'] > class_data['校化学排名'] + 100].copy()
            upper_triangle['区域'] = '其他差化学好'
            
            lower_triangle = class_data[class_data['校总分排名'] < class_data['校化学排名'] - 100].copy()
            lower_triangle['区域'] = '其他好化学差'
            
            class_triangle_data = pd.concat([upper_triangle, lower_triangle])
            class_triangle_data = class_triangle_data.sort_values('校总分排名')
            
            # 添加标题
            worksheet.cell(row=start_row, column=1, value=f'班级{class_num}三角形区域内的学生数据分析')
            
            # 将数据转换为列表并直接写入worksheet
            headers = list(class_triangle_data.columns)
            for col_idx, header in enumerate(headers, 1):
                worksheet.cell(row=start_row + 2, column=col_idx, value=header)
            
            for row_idx, row in enumerate(class_triangle_data.values, 1):
                for col_idx, value in enumerate(row, 1):
                    worksheet.cell(row=start_row + 2 + row_idx, column=col_idx, value=value)

        writer.save()
        writer.close()

    def analyze_all(self, save_plots=True, export_excel=True):
        """分析所有数据"""
        # 第一步：生成所有图片
        self.generate_all_plots()
        
        # 第二步：导出Excel报告
        if export_excel:
            self.export_analysis_to_excel()

def main():
    analyzer = ScoreAnalyzer()
    analyzer.load_data('score.xlsx')
    
    # 可以通过参数控制是否生成独立图片文件和Excel报告
    analyzer.analyze_all(save_plots=True, export_excel=True)

if __name__ == "__main__":
    main()




